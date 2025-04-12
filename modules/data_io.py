# Standard Library Imports
import os
import csv
from collections import Counter, defaultdict
from operator import itemgetter

# Third-Party Library Imports
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import pandas as pd

# Local Application Imports
import modules.config as config

### Utility ###

def parse_products_string(products_string):
    return [
        [name, int(units), int(price)]
        for entry in products_string.split("|") if entry
        for name, units, price in [entry.split(":")]
    ]

def format_products_summary(parsed_products):
    return ", ".join(f"{name} ({units})" for name, units, _ in parsed_products)

### CSV ###

def load_or_create_list_csv(file_name, headers):
    if os.path.exists(file_name):
        with open(file_name, mode="r", newline="") as file:
            reader = csv.reader(file)
            data = list(reader)
            print(f"{file_name} imported.")
            return data
    else:
        with open(file_name, mode="w", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            print(f"{file_name} created.")
        return []
    
def load_or_create_set_csv(file_name, headers):
    if os.path.exists(file_name):
        with open(file_name, mode="r", newline="") as file:
            reader = csv.reader(file)
            next(reader, None)
            result = {row[0] for row in reader if row}
            print(f"{file_name} imported.")
            return result
    else:
        with open(file_name, mode="w", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            print(f"{file_name} created.")
        return set()
    
def write_csv(file_name, rows, headers=None):
    with open(file_name, mode="w", newline="") as file:
        writer = csv.writer(file)
        if headers:
            writer.writerow(headers)
        writer.writerows(rows)
    
def append_csv(file_name, data):
    with open(file_name, mode="a", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(data)

### Excel ###

sales_data = []

def build_daily_summary_sheet(ws, sales_data):
    # Setup headers
    ws.append(config.DAILY_SUMMARY_HEADERS)

    # Setup dictionary for incrementing values
    daily_summary_temp = defaultdict(lambda: {
        "total_sales": 0.0,
        "units_sold": 0,
        "real_rate_total": 0.0,
        "deals": 0,
        "products": Counter(),
        "customers": set(),
        "customer_data": defaultdict(lambda: {"sales": 0.0, "units": 0})
    })

    # Build individual sales data
    for row in sales_data[1:]:
        current_day = int(row[0])
        customer_name = row[1]
        units_sold = int(row[2])
        total_sales = float(row[3])
        products_string = row[6]
        
        # Sum up sales data and customer names
        day_summary = daily_summary_temp[current_day]
        day_summary["units_sold"] += units_sold
        day_summary["deals"] += 1
        day_summary["customers"].add(customer_name)
        day_summary["total_sales"] += total_sales

        # Parse and gather product information
        parsed_products = parse_products_string(products_string)
        parsed_products.sort(key=lambda x: x[1], reverse=True)

        for name, units, _ in parsed_products:
            day_summary["products"][name] += units

        # Gather individual customer total sales and units sold
        cust_data = day_summary["customer_data"][customer_name]
        cust_data["sales"] += total_sales
        cust_data["units"] += units_sold

    for day in sorted(daily_summary_temp):
        data = daily_summary_temp[day]

        # Sort customer list by sales amount
        sorted_customers = sorted(
            data["customer_data"].items(),
            key=lambda item: item[1]["sales"],
            reverse=True
        )

        # Format customer list
        formatted_customers = [
            f"{name} (${info['sales']:.0f} / {info['units']} units)"
            for name, info in sorted_customers
        ]
        customer_summary = ", ".join(formatted_customers)

        sorted_products = sorted(data["products"].items(), key=lambda item: item[1], reverse=True)
        product_summary = ", ".join(f"{name} ({units})" for name, units in sorted_products)

        ws.append([
            day,
            round(data["total_sales"], 2),
            data["units_sold"],
            data["deals"],
            product_summary,
            customer_summary,
        ])

    # Set individual column widths
    column_widths = [8, 16, 10, 10, 50, 150]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width
    
    # Style the header row
    for cell in ws[1]:
        cell.font = Font(color=config.FONT_COLOR, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color=config.CELL_COLOR,
            end_color=config.CELL_COLOR,
            fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin", color=config.FONT_COLOR),
            right=Side(style="thin", color=config.FONT_COLOR),
            top=Side(style="thin", color=config.FONT_COLOR),
            bottom=Side(style="thin", color=config.FONT_COLOR),
        )

    # Style the data rows
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            cell.font = Font(color=config.FONT_COLOR)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color=config.CELL_COLOR,
                end_color=config.CELL_COLOR,
                fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color=config.FONT_COLOR),
                right=Side(style="thin", color=config.FONT_COLOR),
                top=Side(style="thin", color=config.FONT_COLOR),
                bottom=Side(style="thin", color=config.FONT_COLOR),
            )
            if i == 1:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

def build_customer_summary_sheet(ws, sales_data):
    ws.append(config.CUSTOMER_SUMMARY_REPORT_HEADERS)

    # Setup dictionary for incrementing values
    customer_summary_temp = defaultdict(lambda: {
        "total_sales": 0.0,
        "units_total": 0,
        "deals": 0,
        "rates": [],
        "units": [],
        "relationship": "",
        "times_of_day": Counter(),
        "locations": Counter()
    })

    # Build individual sales data
    for row in sales_data[1:]:
        customer_name = row[1]
        units_sold = int(row[2])
        total_sales = float(row[3])
        real_rate = float(row[4])
        location = row[7]
        time_of_day = row[8]
        relationship = row[9]

        # Sum up sales data and customer names
        customer = customer_summary_temp[customer_name]
        customer["total_sales"] += total_sales
        customer["units_total"] += units_sold
        customer["deals"] += 1
        customer["rates"].append(real_rate)
        customer["units"].append(units_sold)
        customer["relationship"] = relationship
        customer["times_of_day"][time_of_day] += 1
        customer["locations"][location] += 1

    for name in sorted(customer_summary_temp):
        data = customer_summary_temp[name]
        avg_sale = data["total_sales"] / data["deals"]
        avg_units = data["units_total"] / data["deals"]
        avg_rate = sum(data["rates"]) / len(data["rates"])

        # Sort time of day and location counts
        sorted_times = sorted(data["times_of_day"].items(), key=itemgetter(1), reverse=True)
        sorted_locs = sorted(data["locations"].items(), key=itemgetter(1), reverse=True)
        formatted_times = ", ".join([f"{time} ({count})" for time, count in sorted_times])
        formatted_locs = ", ".join([f"{loc} ({count})" for loc, count in sorted_locs])

        ws.append([
            name,
            round(data["total_sales"], 2),
            data["units_total"],
            data["deals"],
            round(avg_sale, 2),
            round(avg_units, 2),
            round(avg_rate, 2),
            data["relationship"],
            formatted_times,
            formatted_locs
        ])

    # Set individual column widths
    column_widths = [20, 14, 14, 12, 14, 14, 14, 16, 60, 100]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(color=config.FONT_COLOR, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color=config.CELL_COLOR,
            end_color=config.CELL_COLOR,
            fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin", color=config.FONT_COLOR),
            right=Side(style="thin", color=config.FONT_COLOR),
            top=Side(style="thin", color=config.FONT_COLOR),
            bottom=Side(style="thin", color=config.FONT_COLOR),
        )

    # Style the data rows
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            cell.font = Font(color=config.FONT_COLOR)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color=config.CELL_COLOR,
                end_color=config.CELL_COLOR,
                fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color=config.FONT_COLOR),
                right=Side(style="thin", color=config.FONT_COLOR),
                top=Side(style="thin", color=config.FONT_COLOR),
                bottom=Side(style="thin", color=config.FONT_COLOR),
            )
            if i in (1, 4):  # "TOTAL SALES", "AVG SALE"
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            elif i in (5, 6):  # "AVG UNITS", "AVG RATE"
                cell.number_format = '0.00'

def build_product_summary_sheet(ws, product_data):
    ws.append(config.PRODUCT_SUMMARY_REPORT_HEADERS)

    # Build individual product data
    for row in product_data[1:]:
        product_name = row[0]
        materials = row[1]
        timeframe = int(row[2])
        yield_amount = int(row[3])
        sell_price = int(row[4])

        # Parse materials information
        parsed_materials = [
            {
                "name": name,
                "amount": int(amount),
                "price": int(price)
            }
            for entry in materials.split("|")
            for name, amount, price in [entry.split(":")]
        ]

        # Calculate profit margin per unit
        # Sell price - (total materials cost / yield amount)
        materials_cost = sum(m["amount"] * m["price"] for m in parsed_materials)
        materials_cost_per_unit = materials_cost / yield_amount
        profit_per_unit = sell_price - materials_cost_per_unit

        # Calculate profit per batch and per hour
        profit_per_batch = profit_per_unit * yield_amount
        profit_per_hour = profit_per_batch / timeframe if timeframe > 0 else 0

        # Format timeframe
        timeframe_str = f"{timeframe} hour" if timeframe == 1 else f"{timeframe} hours"

        ws.append([
            product_name,
            round(materials_cost_per_unit, 2),
            sell_price,
            yield_amount,
            timeframe_str,
            round(profit_per_unit, 2),
            round(profit_per_batch, 2),
            round(profit_per_hour, 2)
        ])

    # Set individual column widths
    column_widths = [20, 20, 14, 8, 22, 12, 22, 22]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(color=config.FONT_COLOR, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color=config.CELL_COLOR,
            end_color=config.CELL_COLOR,
            fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin", color=config.FONT_COLOR),
            right=Side(style="thin", color=config.FONT_COLOR),
            top=Side(style="thin", color=config.FONT_COLOR),
            bottom=Side(style="thin", color=config.FONT_COLOR),
        )

    # Style the data rows
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            cell.font = Font(color=config.FONT_COLOR)
            cell.fill = PatternFill(
                start_color=config.CELL_COLOR,
                end_color=config.CELL_COLOR,
                fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color=config.FONT_COLOR),
                right=Side(style="thin", color=config.FONT_COLOR),
                top=Side(style="thin", color=config.FONT_COLOR),
                bottom=Side(style="thin", color=config.FONT_COLOR),
            )
            if i in (1, 2, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            else:
                cell.alignment = Alignment(horizontal="center")


def build_raw_data_sheet(ws, sales_data):
    ws.append(config.RAW_DATA_REPORT_HEADERS)

    # Build individual sales data
    for row in sales_data[1:]:
        current_day = row[0]
        customer_name = row[1]
        units_sold = row[2]
        total_sales = row[3]
        real_rate = row[4]
        ask_rate = row[5]
        products = row[6]
        location = row[7]
        time_of_day = row[8]
        relationship_level = row[9]

        parsed_products = parse_products_string(products)
        parsed_products.sort(key=lambda x: x[1], reverse=True)
        product_summary = ", ".join(f"{name} ({units})" for name, units, _ in parsed_products)

        ws.append([
            int(current_day),
            customer_name,
            int(units_sold),
            float(total_sales),
            float(real_rate),
            float(ask_rate),
            product_summary,
            location,
            time_of_day,
            relationship_level
        ])

    # Set individual column widths
    column_widths = [8, 20, 14, 18, 14, 12, 50, 24, 16, 20]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(color=config.FONT_COLOR, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color=config.CELL_COLOR,
            end_color=config.CELL_COLOR,
            fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin", color=config.FONT_COLOR),
            right=Side(style="thin", color=config.FONT_COLOR),
            top=Side(style="thin", color=config.FONT_COLOR),
            bottom=Side(style="thin", color=config.FONT_COLOR),
        )

    # Style the data rows
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            cell.font = Font(color=config.FONT_COLOR)
            cell.fill = PatternFill(
                start_color=config.CELL_COLOR,
                end_color=config.CELL_COLOR,
                fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color=config.FONT_COLOR),
                right=Side(style="thin", color=config.FONT_COLOR),
                top=Side(style="thin", color=config.FONT_COLOR),
                bottom=Side(style="thin", color=config.FONT_COLOR),
            )
            if i in (0, 2, 6, 7, 8, 9):
                cell.alignment = Alignment(horizontal="center")
            elif i in (3, 4, 5):
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

def export_spreadsheet():
    wb = Workbook()
    ws = wb.active
    sales_data = load_or_create_list_csv(config.SALES_DATA_CSV, config.RAW_DATA_REPORT_HEADERS)
    product_data = load_or_create_list_csv(config.PRODUCT_DATA_CSV, config.PRODUCT_DATA_HEADERS)


    ws.title = config.DAILY_SUMMARY_REPORT_NAME
    build_daily_summary_sheet(ws, sales_data)

    customer_data_ws = wb.create_sheet(title=config.CUSTOMER_SUMMARY_REPORT_NAME)
    build_customer_summary_sheet(customer_data_ws, sales_data)

    product_data_ws = wb.create_sheet(title=config.PRODUCT_SUMMARY_REPORT_NAME)
    build_product_summary_sheet(product_data_ws, product_data)

    raw_data_ws = wb.create_sheet(title=config.RAW_DATA_REPORT_NAME)
    build_raw_data_sheet(raw_data_ws, sales_data)

    os.makedirs("csv", exist_ok=True)

    # Save workbook
    while True:
        try:
            wb.save(config.BUSINESS_REPORT)
            print(f"{config.BUSINESS_REPORT} exported successfully.")
            break
        except PermissionError:
            print(f"Unable to save {config.BUSINESS_REPORT}. Check to see if it is open in another program.")
            while True:
                choice = input("Try again? (y/n): ").strip().lower()
                if choice == "y":
                    print("Retrying...")
                    break
                elif choice == "n":
                    print("Export canceled.")
                    return
                else:
                    print("Please enter 'y' for yes or 'n' for no.")

### Matplotlib ###

def plot_column_over_days(df, column_name, title, y_label, output_filename, color="tab:blue"):
    plt.figure(figsize=(10, 6))
    plt.plot(df["DAY"], df[column_name], marker="o", linestyle="-", color=color)
    plt.title(title)
    plt.xlabel("Day")
    plt.ylabel(y_label)
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(output_filename)
    plt.close()
    print(f"{output_filename} exported successfully.")

def export_figures():
    df = pd.read_excel(config.BUSINESS_REPORT, sheet_name=config.DAILY_SUMMARY_REPORT_NAME, engine="openpyxl")
    os.makedirs("figures", exist_ok=True)

    plot_column_over_days(
        df,
        column_name="TOTAL SALES",
        title="Daily Sales Totals",
        y_label="Sales ($)",
        output_filename=config.SALES_TOTALS_PER_DAY,
        color="tab:blue"
    )

    plot_column_over_days(
        df,
        column_name="UNITS",
        title="Units Sold Per Day",
        y_label="Units",
        output_filename=config.UNITS_SOLD_PER_DAY,
        color="tab:green"
    )

    plot_column_over_days(
        df,
        column_name="DEALS",
        title="Number of Deals Per Day",
        y_label="Number of Deals",
        output_filename=config.DEALS_PER_DAY,
        color="tab:orange"
    )
