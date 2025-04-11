# Standard Library Imports
import os
import csv
from collections import defaultdict

# Third-Party Library Imports
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import pandas as pd

# Local Application Imports
import config

### CSV ###

def load_or_create_list_csv(file_name, headers):
    if os.path.exists(file_name):
        with open(file_name, mode="r", newline="") as file:
            reader = csv.reader(file)
            data = list(reader)
            print(f"{file_name} imported.")
            return data
    else:
        with open(file_name, mode="w", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            print(f"{file_name} created.")
        return []
    
def load_or_create_set_csv(file_name, headers):
    if os.path.exists(file_name):
        with open(file_name, mode="r", newline="") as file:
            reader = csv.reader(file)
            next(reader, None)
            result = {row[0] for row in reader if row}
            print(f"{file_name} imported.")
            return result
    else:
        with open(file_name, mode="w", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            print(f"{file_name} created.")
        return set()
    
def append_csv(file_name, data):
    with open(file_name, mode="a", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(data)

### Excel ###

sales_data = []

def build_daily_summary_sheet(ws, sales_data):
    # Setup headers
    ws.append(config.DAILY_SUMMARY_HEADERS)

    # Setup dictionary for incremental values
    daily_summary_temp = defaultdict(lambda: {
        "total_sales": 0.0,
        "units_sold": 0,
        "real_rate_total": 0.0,
        "ask_rate_total": 0.0,
        "deals": 0,
        "customers": set(),
        "customer_data": defaultdict(lambda: {"sales": 0.0, "units": 0})
    })

    # Build individual sales data
    for row in sales_data[1:]:
        current_day = int(row[0])
        customer_name = row[1]
        units_sold = int(row[2])
        total_sales = float(row[3])
        real_rate = float(row[4])
        ask_rate = float(row[5])
        
        # Sum up sales data and customer names
        daily_summary_temp[current_day]["total_sales"] += total_sales
        daily_summary_temp[current_day]["units_sold"] += units_sold
        daily_summary_temp[current_day]["real_rate_total"] += real_rate
        daily_summary_temp[current_day]["ask_rate_total"] += ask_rate
        daily_summary_temp[current_day]["deals"] += 1
        daily_summary_temp[current_day]["customers"].add(customer_name)

        # Gather individual customer total sales and units sold
        cust_data = daily_summary_temp[current_day]["customer_data"][customer_name]
        cust_data["sales"] += total_sales
        cust_data["units"] += units_sold

    for day in sorted(daily_summary_temp):
        data = daily_summary_temp[day]

        # Calculate average rates
        avg_real_rate = data["real_rate_total"] / data["deals"]
        avg_ask_rate = data["ask_rate_total"] / data["deals"]

        # Sort customer list by sales amount
        sorted_customers = sorted(
            data["customer_data"].items(),
            key=lambda item: item[1]["sales"],
            reverse=True
        )

        # Format customer list
        formatted_customers = [
            f"{name} (${info['sales']:.0f} / {info['units']} units)"
            for name, info in sorted_customers
        ]
        customer_summary = ", ".join(formatted_customers)

        ws.append([
            day,
            round(data["total_sales"], 2),
            data["units_sold"],
            round(avg_real_rate, 2),
            round(avg_ask_rate, 2),
            data["deals"],
            customer_summary
        ])

    # Set individual column widths
    column_widths = [8, 16, 16, 24, 16, 14, 150]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width
    
    # Style the header row
    for cell in ws[1]:
        cell.font = Font(color=config.FONT_COLOR, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color=config.CELL_COLOR,
            end_color=config.CELL_COLOR,
            fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin", color=config.FONT_COLOR),
            right=Side(style="thin", color=config.FONT_COLOR),
            top=Side(style="thin", color=config.FONT_COLOR),
            bottom=Side(style="thin", color=config.FONT_COLOR),
        )

    # Style the data rows
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            cell.font = Font(color=config.FONT_COLOR)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color=config.CELL_COLOR,
                end_color=config.CELL_COLOR,
                fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color=config.FONT_COLOR),
                right=Side(style="thin", color=config.FONT_COLOR),
                top=Side(style="thin", color=config.FONT_COLOR),
                bottom=Side(style="thin", color=config.FONT_COLOR),
            )
            if i in (1, 3, 4):
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

def build_raw_data_sheet(ws, sales_data):
    ws.append(config.CUSTOMER_SUMMARY_HEADERS)

    # Build individual sales data
    for row in sales_data[1:]:
        current_day = row[0]
        customer_name = row[1]
        units_sold = row[2]
        total_sales = row[3]
        real_rate = row[4]
        ask_rate = row[5]
        location = row[6]
        time_of_day = row[7]

        ws.append([
            int(current_day),
            customer_name,
            int(units_sold),
            float(total_sales),
            float(real_rate),
            float(ask_rate),
            location,
            time_of_day
        ])

    # Set individual column widths
    column_widths = [8, 20, 16, 16, 14, 14, 16, 16]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(color=config.FONT_COLOR, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color=config.CELL_COLOR,
            end_color=config.CELL_COLOR,
            fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin", color=config.FONT_COLOR),
            right=Side(style="thin", color=config.FONT_COLOR),
            top=Side(style="thin", color=config.FONT_COLOR),
            bottom=Side(style="thin", color=config.FONT_COLOR),
        )

    # Style the data rows
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            cell.font = Font(color=config.FONT_COLOR)
            cell.fill = PatternFill(
                start_color=config.CELL_COLOR,
                end_color=config.CELL_COLOR,
                fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin", color=config.FONT_COLOR),
                right=Side(style="thin", color=config.FONT_COLOR),
                top=Side(style="thin", color=config.FONT_COLOR),
                bottom=Side(style="thin", color=config.FONT_COLOR),
            )
            if i in (0, 2):
                cell.alignment = Alignment(horizontal="center")
            elif i in (3, 4, 5):
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

def export_spreadsheet():
    wb = Workbook()
    ws = wb.active
    sales_data = load_or_create_list_csv(config.SALES_DATA_CSV, config.CUSTOMER_SUMMARY_HEADERS)

    ws.title = config.DAILY_SUMMARY_REPORT_NAME
    build_daily_summary_sheet(ws, sales_data)

    raw_data_ws = wb.create_sheet(title=config.RAW_DATA_REPORT_NAME)
    build_raw_data_sheet(raw_data_ws, sales_data)

    os.makedirs("csv", exist_ok=True)

    # Save workbook
    while True:
        try:
            wb.save(config.BUSINESS_REPORT)
            print(f"{config.BUSINESS_REPORT} exported successfully.")
            break
        except PermissionError:
            print(f"Unable to save {config.BUSINESS_REPORT}. Check to see if it is open in another program.")
            while True:
                choice = input("Try again? (y/n): ").strip().lower()
                if choice == "y":
                    print("Retrying...")
                    break
                elif choice == "n":
                    print("Export canceled.")
                    return
                else:
                    print("Please enter 'y' for yes or 'n' for no.")

### Matplotlib ###

def plot_column_over_days(df, column_name, title, y_label, output_filename, color="tab:blue"):
    plt.figure(figsize=(10, 6))
    plt.plot(df["DAY"], df[column_name], marker="o", linestyle="-", color=color)
    plt.title(title)
    plt.xlabel("Day")
    plt.ylabel(y_label)
    plt.grid(True)
    plt.tight_layout()
    os.makedirs("figures", exist_ok=True)
    plt.savefig(f"figures/{output_filename}")
    plt.close()

def export_figures():
    df = pd.read_excel(config.BUSINESS_REPORT, sheet_name=config.DAILY_SUMMARY_REPORT_NAME, engine="openpyxl")

    plot_column_over_days(
        df,
        column_name="TOTAL SALES",
        title="Daily Sales Totals",
        y_label="Sales ($)",
        output_filename="daily_sales_totals.png",
        color="tab:blue"
    )

    plot_column_over_days(
        df,
        column_name="UNITS SOLD",
        title="Units Sold Per Day",
        y_label="Units",
        output_filename="daily_units_sold.png",
        color="tab:green"
    )

    plot_column_over_days(
        df,
        column_name="DEALS",
        title="Number of Customers Per Day",
        y_label="Number of Deals",
        output_filename="daily_deals.png",
        color="tab:orange"
    )
